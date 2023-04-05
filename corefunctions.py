import os
import win32com.client
import datetime


from openpyxl import load_workbook, Workbook

def sendmailaclar():
    from datetime import datetime
    outlook = win32com.client.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    now = datetime.now()
    tod = now.strftime("%d-%m-%Y") 
    today=str(tod)
    email = "ivan.acosta@corusconsulting.com"
    mail.To = email
    mail.Subject = 'Solicitud de aclaracion'
    pathfilealcar="C:/Users/E-IACOSTA/Desktop/RPA/core/Bitacoras/%sBITACORAs.xlsx/Bitacora.xlsx"%today
    mail.Attachments.Add(pathfilealcar)
    mail.HTMLBody = """
                    <html lang="en">
                    <body>
                    <div>
                    <p>Saludos Cordiales</p></br></br></br>
                    <p>Solicitud de Aclaraci&oacuten </p>
                    </div>
                    </b></br></br></br>
                    </body>
                    </body>
                    </html> """
    mail.Send()

def mergebitacora():

        # importing the required modules
    import glob
    import pandas as pd
    # specifying the path to csv files
    path = "C:/Users/E-IACOSTA/Desktop/RPA/core/Bitacoras/"
    # csv files in the path
    file_list = glob.glob(path + "/*.xlsx")
    # list of excel files we want to merge.
    #pd.read_excel(file_path) reads the excel
    # data into pandas dataframe.
    excl_list = []

    for file in file_list:
        excl_list.append(pd.read_excel(file))

    # create a new dataframe to store the
    # merged excel file.
    excl_merged = pd.DataFrame()

    for excl_file in excl_list:
        # appends the data into the excl_merged
        # dataframe.
        excl_merged = excl_merged.append(
        excl_file, ignore_index=True)

    # exports the dataframe into excel file with
    # specified name.
    dirsave="C:/Users/E-IACOSTA/Desktop/RPA/core/BitacoraHistorico/BitacoraHistorica.xlsx"
    excl_merged.to_excel(dirsave, index=False)

def bajarxcel():
  from datetime import datetime
  path = os.path.expanduser("C:/Users/E-IACOSTA/Desktop/RPA/core/AclaracionesWeek/")

  now = datetime.now()
  tod = now.strftime("%d-%m-%Y") 
  today=str(tod)
  #print(today)
  outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
  inbox = outlook.GetDefaultFolder(6)
  messages = inbox.Items
  subject = "Aclaraciones PF"

  for message in messages:
      if message.Subject == subject:
          attNumber = message.Attachments.Count
          if attNumber > 0:
              for n in range(attNumber):
                  att = message.Attachments.Item(n+1)
                  attName =  str(today)+att.DisplayName
                  if not 'image' in attName:
                      #print (attName)
                      att.SaveAsfile(os.path.join(path,attName))

def is_list_empty(dup):
    # checking the length
    if len(dup) == 0:
        return False
    return True

def procesoaclaracion():
    
      import time
      from datetime import timedelta
#######################====Se baja el excell del correo=========##################################################
      bajarxcel()
      time.sleep(5)
########################=======LIBRERIAS=====##############################################
      import os
      import random
      from datetime import  datetime, timedelta
      import openpyxl
################====Desarga se renombra el archivo segun la fecha#######################

      now = datetime.now()
      tt = now.strftime("%d-%m-%Y") 
      ttt=str(tt)
      filesheet = "C:/Users/E-IACOSTA/Desktop/RPA/core/AclaracionesWeek/%sWeekAclaraciones.xlsx"%ttt#Ruta del archivo
      wb_obj = openpyxl.load_workbook(filesheet)
      sheet = wb_obj.active
#CABECERA
      col_names = []
      for column in sheet.iter_rows(sheet.max_row, sheet.max_column):
          col_names.append(column.value)
#COL 1:IDS
      col_names1 = []
      for column in sheet.iter_rows(2, sheet.max_row):
          if column[2].value <=300:
              col_names1.append(column[0].value)
#COL 2:NOMBRES 
      col_names2 = []
      for column in sheet.iter_rows(2, sheet.max_row):
          if column[2].value <=300:
              col_names2.append(column[1].value)
#COL 3:MONTOS  
      col_names3 = []
      for column in sheet.iter_rows(2, sheet.max_row):
          if column[2].value <=300:
              col_names3.append(column[2].value)
#COL 4:FECHAS  
      col_names4 = []
      for column in sheet.iter_rows(2, sheet.max_row):
          if column[2].value <=300:
              col_names4.append(column[3].value)
              
      colsum1=[col_names1,col_names2,col_names3,col_names4]
#################################=GENERA BITACORA=##########################################################################
      import pandas as pd
      from pandas import ExcelWriter
      # decodigo.com
      df = pd.DataFrame({'ID': colsum1[0],
                      'NOMBRE': colsum1[1],
                      'MONTO':colsum1[2],
                      'FECHA':colsum1[3]})
      df = df[['ID', 'NOMBRE', 'MONTO','FECHA']]
      writer = ExcelWriter('C:/Users/E-IACOSTA/Desktop/RPA/core/Bitacoras/%sBITACORA.xlsx'%ttt)
      df.to_excel(writer, 'Hoja de datos', index=False)
      writer.save()
######################################=MERGE BITACORAS =#######################################################
      mergebitacora()
      time.sleep(5)
################# =COMPARA BITACORA HISTORICA CON WEEK= ####################################### 
      filesheetcomp = "C:/Users/E-IACOSTA/Desktop/RPA/core/BitacoraHistorico/BitacoraHistorica.xlsx"#Ruta del archivo
      wb_obj = openpyxl.load_workbook(filesheetcomp)
      sheet = wb_obj.active
#COL 1:IDS
      colids = []
      for column in sheet.iter_rows(2, sheet.max_row):
          colids.append(column[0].value)
#COL 2:NOMBRES 
      colnames = []
      for column in sheet.iter_rows(2, sheet.max_row):
          colnames.append(column[1].value)
#COL 3:MONTOS  
      colmontos = []
      for column in sheet.iter_rows(2, sheet.max_row):
          colmontos.append(column[2].value)
#COL 4:FECHAS  
      colfechas = []
      for column in sheet.iter_rows(2, sheet.max_row): 
          colfechas.append(column[3].value)

      colsum2=[colids,colnames,colmontos,colfechas]
      #print(colsum2[3][-1])
##########################= VALIDA FECHAS =##############################################
      
      from datetime import datetime
      from datetime import timedelta
      fc=colsum2[3][-1]
      fc + timedelta(31,3600)
      hoy=datetime.now() - fc
      hoystr=str(hoy)
      hoyspli=hoystr.split(" ")
      hhh=hoyspli[0]
      hoydiferencia=int(hhh)
      flagtime=0
      if hoydiferencia < 30:
          flagtime=1
          print(flagtime)
      elif hoydiferencia >30:
          flagtime=2
          print(flagtime)
      elif hoydiferencia >180:
            flagtime=3
            print(flagtime)
#########################=====VALIDA REICIDENCIA
      #colnames=lista de bitacora historica
      #colnames2=lista de weekmail
      flagreincide=0
      listareincidente=colnames + col_names2
      visited=set()
      dup=[x for x in listareincidente if x in visited or (visited.add(x) or False)]
      print(dup)
      print(is_list_empty(dup))

      if is_list_empty(dup) == True:
          flagreincide=1
          print(flagreincide)
          #Enviar correocon lista de reincidentes
      elif is_list_empty(dup)== False:
          flagreincide=2
          print(flagreincide)
          #No hay reincidentes
####################################=VALIDA FECHAS Y REINCIDENCIAS=#############################################
      flagmails=0
      #Reincidencia menor a un mes
      if flagreincide == 2 & flagtime ==1:
          flagmails=1
          print(flagmails,"No reincide y menor a un mes")
       #Reincide mayor a un mes
      elif flagreincide == 1 & flagtime ==2:
          flagmails=2
          print(flagmails,"Reinicide mayor a un mes")
        #Reincide en 6 meses
      elif flagreincide == 1 & flagtime ==3:
          flagmails=3
          print(flagmails,"Reincide en 6 meses ")
      elif flagreincide == 1 & flagtime ==1:
          flagmails=4
          print(flagmails,"Reincide menos de un mes ")

####################################=CASES ACTIONS =#############################################

      match flagmails:
          case 1:
              print("Se envia correo de aclaracion")
          case 2:
              print("Se envia correo de verficacion ya que reincide")
          case 3:
              print("Se solicita bloqueo ya que reincide en lapso de 6 meses ")
          case 4:
              print("Se envia correo de verificacion ya que reincide en menos de un mes")